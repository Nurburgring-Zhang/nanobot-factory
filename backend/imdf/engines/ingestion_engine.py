"""数据导入引擎-支持CSV/JSON/Excel格式

P21 Phase 2 P1 critical fix (R2-NEW-#2 + R2-NEW-#3):

* **R2-NEW-#2 (id column collision)** — 之前的 `_insert_rows` 把 PK 硬编码为
  ``id INTEGER PRIMARY KEY AUTOINCREMENT`` 并追加 CSV 的列定义, 当 CSV 自带
  ``id`` 列时 SQLite 抛 ``duplicate column name: id``。修复: 检查 CSV 表头是否
  已经包含 ``id`` (大小写不敏感), 若是, 把 PK 重命名为 ``row_id`` 并把用户
  的 ``id`` 当作普通 TEXT 列导入。

* **R2-NEW-#3 (inconsistent-row silent loss)** — ``csv.DictReader`` 对列数不
  一致的行只把缺失字段填成 ``None``, 而旧逻辑用 ``str(row.get(c, ""))`` 把
  ``None`` 序列化成字符串 ``"None"`` 写进 DB, 出现静默数据损坏。修复: 用
  ``csv.reader`` (而非 DictReader) 先做严格表头对齐检查, 任何行
  ``len(row) != len(header)`` 直接抛 :class:`IngestionError`, 整个导入
  rollback。

向后兼容:
  * 没有 ``id`` 列的 CSV — PK 仍是 ``id`` (与旧行为一致)。
  * 现有 ``import_json`` / ``import_excel`` 路径不动 (Excel 末尾空值在 read
    模式下常见, 这里不抛错)。
  * 返回 dict 增加 ``pk_column`` 字段, 旧代码读 ``rows_imported`` 不会被破坏。
"""
import csv
import json
import os
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger(__name__)


class IngestionError(Exception):
    """Ingestion 引擎在数据完整性无法保证时抛此异常。

    使用场景:
      * CSV 行 / 表头列数不一致 (R2-NEW-#3)
      * CSV / JSON 字段名包含 SQLite 保留字且无法安全转义
      * 任何结构性问题 — engine 不再做静默修复
    """


class IngestionEngine:
    # PK 名字优先级: 用户列里已经有 'id' (大小写不敏感) 时让位, 避免碰撞
    _USER_ID_LIKE = "id"
    _FALLBACK_PK = "row_id"
    # _imported_at 是 engine 内部 audit 列, 同样需要避让
    _RESERVED_COLUMNS = frozenset({"_imported_at"})

    def __init__(self, db_path: str = None):
        self.db_path = db_path or os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "..", "data", "imdf.db"
        )

    # ------------------------------------------------------------------ CSV
    def import_csv(self, file_path: str, table: str = "imported_data") -> dict:
        """导入 CSV 到 SQLite 表。

        Args:
            file_path: CSV 文件路径 (UTF-8)。
            table: 目标 SQLite 表名 (建议每次 import 用新表名, 避免
                   ``CREATE TABLE IF NOT EXISTS`` 的 schema 锁)。

        Returns:
            dict — ``{"success": True, "data": {"table", "rows_imported",
            "columns", "total_in_file", "pk_column"}}`` 或
            ``{"success": False, "error": "..."}``。

        Raises:
            IngestionError: CSV 中存在行/表头列数不一致 (R2-NEW-#3)。
            FileNotFoundError: 通过 ``{"success": False, "error": ...}`` 返回,
                不抛异常 (向后兼容)。
        """
        if not os.path.exists(file_path):
            return {"success": False, "error": f"文件不存在: {file_path}"}
        # R2-NEW-#3: 用 csv.reader 严格校验, 再转 dict 喂给 _insert_rows。
        # 任何行 len(row) != len(header) 直接抛 IngestionError, 整个导入不
        # 写 DB。
        try:
            header, data_rows = self._read_csv_strict(file_path)
        except IngestionError:
            raise
        if header is None:
            return {"success": False, "error": f"文件不存在: {file_path}"}
        if not header:
            return {"success": False, "error": "空文件"}
        if not data_rows:
            return {"success": False, "error": "空文件 (无数据行)"}
        # 列名去重 — DictReader 在 header 含重复列名时静默取最后一个, 这里
        # 显式抛错, 避免后续 SQLite INSERT 报 ambiguous column。
        seen = set()
        for c in header:
            if c in seen:
                raise IngestionError(
                    f"CSV 表头含重复列名: '{c}' — 请先规范化表头"
                )
            seen.add(c)
        # Reserved column collision: 表头里出现 _imported_at 会让 INSERT 报
        # ambiguous column。把用户的 ``_imported_at`` 重命名为 ``user_imported_at``
        # (去掉前导下划线再加 ``user_`` 前缀 — 避免 ``user___imported_at`` 这种
        # 三下划线难看的名字)。
        normalized_header = [
            ("user_" + c.lstrip("_")) if c in self._RESERVED_COLUMNS else c
            for c in header
        ]
        rows = [
            {normalized_header[i]: r[i] if i < len(r) else "" for i in range(len(normalized_header))}
            for r in data_rows
        ]
        return self._insert_rows(rows, table, header=normalized_header)

    @staticmethod
    def _read_csv_strict(file_path: str):
        """严格读取 CSV — 任何行列数 != 表头列数都抛 IngestionError。

        Returns:
            (header, data_rows) — header 是 list[str], data_rows 是
            list[list[str]]。文件不存在时返回 (None, None)。
        """
        if not os.path.exists(file_path):
            return None, None
        with open(file_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            all_rows = list(reader)
        if not all_rows:
            return [], []
        header = all_rows[0]
        expected = len(header)
        # 检查 header 自身 — 空 header 也算非法 (len==0 时任何数据行都通不过)
        if expected == 0:
            raise IngestionError("CSV 表头为空 — 请检查文件格式")
        for idx, row in enumerate(all_rows[1:], start=2):  # row 1 = header
            if len(row) != expected:
                raise IngestionError(
                    f"行 {idx} 有 {len(row)} 列, 表头有 {expected} 列 "
                    f"(表头: {header}, 该行: {row})"
                )
        return header, all_rows[1:]

    # ----------------------------------------------------------------- JSON
    def import_json(self, file_path: str, table: str = "imported_data") -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"文件不存在: {file_path}"}
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            data = [data]
        return self._insert_rows(data, table)

    # ----------------------------------------------------------------- Excel
    def import_excel(self, file_path: str, table: str = "imported_data") -> dict:
        try:
            import openpyxl
        except ImportError:
            return {"success": False, "error": "需要安装openpyxl: pip install openpyxl"}
        if not os.path.exists(file_path):
            return {"success": False, "error": f"文件不存在: {file_path}"}
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Excel 允许表头单元格为 None; 显式 rename 为 col_N, 否则后续 dict
        # 用 None 作 key 会引发 TypeError。
        headers = [
            (f"col_{i}" if h is None or str(h).strip() == "" else str(h))
            for i, h in enumerate(raw_headers)
        ]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Excel read_only 模式常见 trailing None, 静默补齐到表头长度。
            row_list = list(row) if row is not None else []
            if len(row_list) < len(headers):
                row_list = row_list + [None] * (len(headers) - len(row_list))
            rows.append({headers[i]: row_list[i] for i in range(len(headers))})
        return self._insert_rows(rows, table, header=headers)

    # ----------------------------------------------------------------- core
    def _insert_rows(
        self,
        rows: List[Dict],
        table: str,
        header: Optional[List[str]] = None,
    ) -> dict:
        """通用 INSERT — JSON / Excel / CSV 都走这里。

        Args:
            rows: 已经是 list[dict], key 与 header 一一对应。
            table: SQLite 表名。
            header: 列名顺序。若为 None, 用 rows[0].keys() (旧行为)。
        """
        if not rows:
            return {"success": False, "error": "无数据"}
        if header is None:
            header = list(rows[0].keys())
        cols = list(header)
        # R2-NEW-#2: 决定 PK 名 — 用户已含 'id' (大小写不敏感) 时让位。
        pk_col = self._FALLBACK_PK if any(
            c.lower() == self._USER_ID_LIKE for c in cols
        ) else self._USER_ID_LIKE
        # 防御: 用户的列名里如果出现与 PK 同名 (row_id), 同样让位。极端情况。
        if pk_col in {c.lower() for c in cols}:
            pk_col = "_ingest_pk"
        col_defs = ", ".join([f'"{c}" TEXT' for c in cols])
        conn = sqlite3.connect(self.db_path)
        try:
            cursor = conn.cursor()
            cursor.execute(
                f'CREATE TABLE IF NOT EXISTS [{table}] '
                f'({pk_col} INTEGER PRIMARY KEY AUTOINCREMENT, _imported_at TEXT, {col_defs})'
            )
            ts = datetime.now().isoformat()
            cols_str = ", ".join([f'"{c}"' for c in cols])
            # 占位符 = 列数 + 1 (含 _imported_at)
            vals_str = ", ".join(["?"] * (len(cols) + 1))
            inserted = 0
            # R2-NEW-#3: 任何行级 DB 错误都不再静默吃掉 — 让 caller 拿到异常。
            for row in rows:
                values = []
                for c in cols:
                    v = row.get(c) if hasattr(row, "get") else None
                    if v is None:
                        values.append("")
                    else:
                        values.append(str(v))
                values.append(ts)
                cursor.execute(
                    f'INSERT INTO [{table}] ({cols_str}, _imported_at) '
                    f'VALUES ({vals_str})',
                    values,
                )
                inserted += 1
            conn.commit()
        except Exception:
            # 任何 DB 错误 — 强制 rollback + 重新抛出, 避免 half-committed
            # 状态留在 imdf.db 里。
            try:
                conn.rollback()
            except Exception:
                pass
            raise
        finally:
            conn.close()
        return {
            "success": True,
            "data": {
                "table": table,
                "rows_imported": inserted,
                "columns": cols,
                "total_in_file": len(rows),
                "pk_column": pk_col,
            },
        }
